"""
Generate Excel Dashboard for Trading Analytics
"""

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, LineChart

# ============================================================================
# DATA LOADING (same as analysis script)
# ============================================================================

DATA_PATH = Path(__file__).parent.parent / 'data'

def load_and_transform():
    trades = pd.read_csv(DATA_PATH / 'trades_raw.csv', parse_dates=['open_time', 'close_time'])
    accounts = pd.read_csv(DATA_PATH / 'accounts_raw.csv', parse_dates=['created_at', 'closed_at'])
    clients = pd.read_csv(DATA_PATH / 'clients_raw.csv', parse_dates=['created_at'])
    balances = pd.read_csv(DATA_PATH / 'balances_eod_raw.csv', parse_dates=['date'])
    symbols = pd.read_csv(DATA_PATH / 'symbols_ref.csv')
    
    trades['side'] = trades['side'].str.upper().replace({'B': 'BUY', 'S': 'SELL'})
    clients['segment'] = clients['segment'].str.lower().str.strip()
    trades['trade_date'] = trades['open_time'].dt.date
    trades['trade_week'] = pd.to_datetime(trades['trade_date']).dt.to_period('W').dt.start_time
    
    spine = accounts.merge(clients, on='client_id', how='left', suffixes=('_acct', '_client'))
    
    trades_enriched = trades.merge(
        spine[['account_id', 'platform', 'client_id', 'segment', 'jurisdiction', 'is_system', 'is_deleted']],
        on=['account_id', 'platform'], how='left'
    )
    
    trades_enriched = trades_enriched.merge(
        symbols[['platform', 'platform_symbol', 'std_symbol', 'asset_class']],
        left_on=['platform', 'symbol'], right_on=['platform', 'platform_symbol'], how='left'
    )
    trades_enriched['std_symbol'] = trades_enriched['std_symbol'].fillna(trades_enriched['symbol'])
    trades_enriched['net_pnl'] = trades_enriched['realized_pnl'] + trades_enriched['commission']
    trades_enriched['is_cancelled'] = trades_enriched['status'] == 'CANCELLED'
    
    analysis_trades = trades_enriched[
        (~trades_enriched['is_cancelled']) & (trades_enriched['is_system'] != True)
    ].copy()
    
    balances = balances.sort_values(['account_id', 'platform', 'date'])
    balances['peak_equity'] = balances.groupby(['account_id', 'platform'])['equity'].cummax()
    balances['drawdown_pct'] = (balances['equity'] - balances['peak_equity']) / balances['peak_equity']
    
    return {'trades': analysis_trades, 'trades_all': trades_enriched, 'balances': balances, 'spine': spine}

# ============================================================================
# STYLING
# ============================================================================

HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(bold=True, size=14, color='1F4E79')
NUMBER_FORMAT = '#,##0.00'
PCT_FORMAT = '0.0%'
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER

def add_df_to_sheet(ws, df, start_row, title=None):
    if title:
        ws.cell(row=start_row, column=1, value=title).font = TITLE_FONT
        start_row += 1
    
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=start_row + r_idx, column=c_idx, value=value)
            cell.border = THIN_BORDER
            if r_idx == 0:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal='center')
            elif isinstance(value, (int, float)) and not pd.isna(value):
                cell.number_format = NUMBER_FORMAT
    
    return start_row + len(df) + 2

# ============================================================================
# DASHBOARD GENERATION
# ============================================================================

def create_dashboard():
    data = load_and_transform()
    trades = data['trades']
    
    wb = Workbook()
    
    # === EXECUTIVE SUMMARY ===
    ws = wb.active
    ws.title = 'Executive Summary'
    
    ws['A1'] = 'TRADING ANALYTICS DASHBOARD'
    ws['A1'].font = Font(bold=True, size=18, color='1F4E79')
    ws['A2'] = f'Data Period: {trades["trade_date"].min()} to {trades["trade_date"].max()}'
    
    # KPIs
    ws['A4'] = 'KEY METRICS'
    ws['A4'].font = TITLE_FONT
    
    kpis = [
        ('Total Net PnL', trades['net_pnl'].sum()),
        ('Total Trades', len(trades)),
        ('Total Volume (lots)', trades['volume'].sum()),
        ('Avg PnL per Trade', trades['net_pnl'].mean()),
        ('Win Rate', (trades['net_pnl'] > 0).mean()),
        ('Active Clients', trades['client_id'].nunique()),
        ('Active Accounts', trades['account_id'].nunique()),
    ]
    
    for i, (label, value) in enumerate(kpis):
        ws.cell(row=5+i, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=5+i, column=2, value=value)
        if 'Rate' in label:
            cell.number_format = PCT_FORMAT
        else:
            cell.number_format = NUMBER_FORMAT
    
    # Segment summary
    seg_df = trades.groupby('segment').agg(
        Clients=('client_id', 'nunique'),
        Trades=('trade_id', 'count'),
        Net_PnL=('net_pnl', 'sum'),
        Avg_PnL=('net_pnl', 'mean')
    ).reset_index()
    seg_df.columns = ['Segment', 'Clients', 'Trades', 'Net PnL', 'Avg PnL']
    add_df_to_sheet(ws, seg_df, 14, 'PERFORMANCE BY SEGMENT')
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    
    # === TOP/BOTTOM PERFORMERS ===
    ws2 = wb.create_sheet('Performance')
    
    client_perf = trades.groupby(['client_id', 'segment']).agg(
        Trades=('trade_id', 'count'),
        Volume=('volume', 'sum'),
        Net_PnL=('net_pnl', 'sum')
    ).reset_index()
    client_perf.columns = ['Client ID', 'Segment', 'Trades', 'Volume', 'Net PnL']
    
    top_10 = client_perf.nlargest(10, 'Net PnL')
    bottom_10 = client_perf.nsmallest(10, 'Net PnL')
    
    row = add_df_to_sheet(ws2, top_10, 1, '🏆 TOP 10 PERFORMERS')
    add_df_to_sheet(ws2, bottom_10, row, '⚠️ BOTTOM 10 PERFORMERS')
    
    # Loss symbols
    symbol_perf = trades.groupby('std_symbol').agg(
        Trades=('trade_id', 'count'),
        Net_PnL=('net_pnl', 'sum')
    ).reset_index()
    symbol_perf.columns = ['Symbol', 'Trades', 'Net PnL']
    loss_symbols = symbol_perf[symbol_perf['Net PnL'] < 0].nsmallest(10, 'Net PnL')
    
    add_df_to_sheet(ws2, loss_symbols, row + len(bottom_10) + 4, '📉 LOSS-DRIVING SYMBOLS')
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws2.column_dimensions[col].width = 15
    
    # === ACTIVITY ===
    ws3 = wb.create_sheet('Activity')
    
    daily = trades.groupby('trade_date').agg(
        Active_Clients=('client_id', 'nunique'),
        Active_Accounts=('account_id', 'nunique'),
        Trades=('trade_id', 'count')
    ).reset_index()
    daily.columns = ['Date', 'Active Clients', 'Active Accounts', 'Trades']
    
    add_df_to_sheet(ws3, daily, 1, '📅 DAILY ACTIVITY')
    
    # Trade size by segment
    size_seg = trades.groupby('segment').agg(
        Avg_Volume=('volume', 'mean'),
        Median_Volume=('volume', 'median'),
        Trades=('trade_id', 'count')
    ).reset_index()
    size_seg.columns = ['Segment', 'Avg Volume', 'Median Volume', 'Trades']
    
    add_df_to_sheet(ws3, size_seg, len(daily) + 5, '📊 TRADE SIZE BY SEGMENT')
    
    for col in ['A', 'B', 'C', 'D']:
        ws3.column_dimensions[col].width = 18
    
    # === RISK ===
    ws4 = wb.create_sheet('Risk')
    
    balances = data['balances']
    worst_dd = balances.groupby(['account_id', 'platform']).agg(
        Min_Drawdown=('drawdown_pct', 'min'),
        Current_Equity=('equity', 'last'),
        Peak_Equity=('peak_equity', 'max')
    ).reset_index()
    worst_dd = worst_dd.merge(
        data['spine'][['account_id', 'platform', 'client_id']].drop_duplicates(),
        on=['account_id', 'platform'], how='left'
    )
    worst_dd = worst_dd.nsmallest(10, 'Min_Drawdown')[['account_id', 'client_id', 'Min_Drawdown', 'Current_Equity', 'Peak_Equity']]
    worst_dd.columns = ['Account', 'Client', 'Max Drawdown %', 'Current Equity', 'Peak Equity']
    
    row = add_df_to_sheet(ws4, worst_dd, 1, '🔴 LARGEST DRAWDOWNS')
    
    # Deleted accounts trading
    deleted = data['trades_all'][
        (data['trades_all']['is_deleted'] == True) & 
        (~data['trades_all']['is_cancelled'])
    ].groupby(['account_id', 'client_id']).agg(
        Trades=('trade_id', 'count'),
        Last_Trade=('trade_date', 'max'),
        Net_PnL=('net_pnl', 'sum')
    ).reset_index()
    deleted.columns = ['Account', 'Client', 'Trades', 'Last Trade', 'Net PnL']
    
    add_df_to_sheet(ws4, deleted, row, '⚠️ DELETED ACCOUNTS STILL TRADING')
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws4.column_dimensions[col].width = 18
    
    # Save
    output_path = Path(__file__).parent.parent / 'trading_dashboard.xlsx'
    wb.save(output_path)
    print(f"Dashboard saved to: {output_path}")
    return output_path

if __name__ == '__main__':
    create_dashboard()
